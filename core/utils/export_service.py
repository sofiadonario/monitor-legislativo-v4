"""
Export service for various output formats
"""

import os
import csv
import json
from datetime import datetime
from typing import List, Optional
import logging

from ..models.models import Proposition, SearchResult


class ExportService:
    """Service for exporting search results to various formats"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def export(self, results: List[SearchResult], format: str, 
              output_path: str, metadata: Optional[dict] = None) -> bool:
        """
        Export search results to specified format
        
        Args:
            results: List of SearchResult objects
            format: Export format (CSV, HTML, PDF, JSON, XLSX)
            output_path: Path to save the exported file
            metadata: Additional metadata to include
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            format_upper = format.upper()
            
            if format_upper == "CSV":
                return self._export_csv(results, output_path, metadata)
            elif format_upper == "HTML":
                return self._export_html(results, output_path, metadata)
            elif format_upper == "PDF":
                return self._export_pdf(results, output_path, metadata)
            elif format_upper == "JSON":
                return self._export_json(results, output_path, metadata)
            elif format_upper == "XLSX":
                return self._export_xlsx(results, output_path, metadata)
            else:
                self.logger.error(f"Unsupported export format: {format}")
                return False
                
        except Exception as e:
            self.logger.error(f"Export failed: {str(e)}")
            return False
    
    def _export_csv(self, results: List[SearchResult], output_path: str, 
                   metadata: Optional[dict] = None) -> bool:
        """Export to CSV format"""
        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = [
                    'Fonte', 'Tipo', 'Número', 'Ano', 'Título', 
                    'Resumo', 'Autores', 'Data Publicação', 'Status', 'URL'
                ]
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for result in results:
                    for prop in result.propositions:
                        writer.writerow({
                            'Fonte': prop.source.value,
                            'Tipo': prop.type.value,
                            'Número': prop.number,
                            'Ano': prop.year,
                            'Título': prop.title,
                            'Resumo': prop.summary[:500],
                            'Autores': prop.author_names,
                            'Data Publicação': prop.publication_date.strftime('%d/%m/%Y'),
                            'Status': prop.status.value,
                            'URL': prop.url
                        })
                
                self.logger.info(f"Exported to CSV: {output_path}")
                return True
                
        except Exception as e:
            self.logger.error(f"CSV export failed: {str(e)}")
            return False
    
    def _export_html(self, results: List[SearchResult], output_path: str,
                    metadata: Optional[dict] = None) -> bool:
        """Export to HTML format"""
        try:
            html_content = self._generate_html(results, metadata)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.info(f"Exported to HTML: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"HTML export failed: {str(e)}")
            return False
    
    def _generate_html(self, results: List[SearchResult], 
                      metadata: Optional[dict] = None) -> str:
        """Generate HTML content"""
        export_date = datetime.now().strftime('%d/%m/%Y %H:%M')
        total_results = sum(len(r.propositions) for r in results)
        
        html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Monitor de Políticas Públicas - Relatório de Busca</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .header {{
            background-color: #003366;
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .header h1 {{
            margin: 0;
            font-size: 24px;
        }}
        .metadata {{
            background-color: white;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .results {{
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background-color: #003366;
            color: white;
            font-weight: bold;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        a {{
            color: #0066CC;
            text-decoration: none;
        }}
        a:hover {{
            text-decoration: underline;
        }}
        .source-section {{
            margin-bottom: 30px;
        }}
        .source-title {{
            background-color: #0066CC;
            color: white;
            padding: 10px;
            border-radius: 4px;
            margin-bottom: 10px;
        }}
        .summary {{
            max-width: 400px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}
        .footer {{
            text-align: center;
            margin-top: 30px;
            color: #666;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Monitor de Políticas Públicas MackIntegridade</h1>
        <p>Relatório de Busca Legislativa</p>
    </div>
    
    <div class="metadata">
        <h2>Informações da Busca</h2>
        <p><strong>Data:</strong> {export_date}</p>
        <p><strong>Total de Resultados:</strong> {total_results}</p>
"""
        
        if metadata:
            if 'query' in metadata:
                html += f"        <p><strong>Termo de Busca:</strong> {metadata['query']}</p>\n"
            if 'filters' in metadata:
                html += f"        <p><strong>Filtros:</strong> {metadata['filters']}</p>\n"
        
        html += """    </div>
    
    <div class="results">
        <h2>Resultados da Busca</h2>
"""
        
        # Group results by source
        for result in results:
            if not result.propositions:
                continue
                
            source_name = result.source.value if result.source else "Fonte Desconhecida"
            html += f"""
        <div class="source-section">
            <div class="source-title">{source_name} ({len(result.propositions)} resultados)</div>
            <table>
                <thead>
                    <tr>
                        <th>Tipo</th>
                        <th>Número/Ano</th>
                        <th>Título</th>
                        <th>Autores</th>
                        <th>Data</th>
                        <th>Ação</th>
                    </tr>
                </thead>
                <tbody>
"""
            
            for prop in result.propositions:
                html += f"""
                    <tr>
                        <td>{prop.type.value}</td>
                        <td>{prop.number}/{prop.year}</td>
                        <td class="summary" title="{prop.summary}">{prop.title}</td>
                        <td>{prop.author_names}</td>
                        <td>{prop.publication_date.strftime('%d/%m/%Y')}</td>
                        <td><a href="{prop.url}" target="_blank">Ver Detalhes</a></td>
                    </tr>
"""
            
            html += """
                </tbody>
            </table>
        </div>
"""
        
        html += """
    </div>
    
    <div class="footer">
        <p>Gerado pelo Monitor de Políticas Públicas MackIntegridade</p>
        <p>© 2025 MackIntegridade - Todos os direitos reservados</p>
    </div>
</body>
</html>
"""
        
        return html
    
    def _export_pdf(self, results: List[SearchResult], output_path: str,
                   metadata: Optional[dict] = None) -> bool:
        """Export to PDF format"""
        try:
            # Try different PDF libraries
            pdf_generated = False
            
            # Try WeasyPrint first
            try:
                from weasyprint import HTML
                html_content = self._generate_html(results, metadata)
                HTML(string=html_content).write_pdf(output_path)
                pdf_generated = True
            except ImportError:
                pass
            
            # Try pdfkit
            if not pdf_generated:
                try:
                    import pdfkit
                    html_content = self._generate_html(results, metadata)
                    pdfkit.from_string(html_content, output_path)
                    pdf_generated = True
                except ImportError:
                    pass
            
            # Try reportlab
            if not pdf_generated:
                try:
                    from reportlab.lib import colors
                    from reportlab.lib.pagesizes import letter, landscape
                    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet
                    
                    doc = SimpleDocTemplate(output_path, pagesize=landscape(letter))
                    elements = []
                    
                    # Title
                    styles = getSampleStyleSheet()
                    elements.append(Paragraph("Monitor de Políticas Públicas - Relatório", styles['Title']))
                    elements.append(Spacer(1, 12))
                    
                    # Data
                    for result in results:
                        if not result.propositions:
                            continue
                        
                        # Source header
                        elements.append(Paragraph(f"{result.source.value}", styles['Heading2']))
                        elements.append(Spacer(1, 6))
                        
                        # Table data
                        data = [['Tipo', 'Número', 'Título', 'Data']]
                        for prop in result.propositions[:20]:  # Limit to 20 per source
                            data.append([
                                prop.type.value,
                                f"{prop.number}/{prop.year}",
                                prop.title[:50] + "...",
                                prop.publication_date.strftime('%d/%m/%Y')
                            ])
                        
                        # Create table
                        t = Table(data)
                        t.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        
                        elements.append(t)
                        elements.append(Spacer(1, 12))
                    
                    doc.build(elements)
                    pdf_generated = True
                    
                except ImportError:
                    pass
            
            if not pdf_generated:
                # Fallback: save as HTML if no PDF library available
                self.logger.warning("No PDF library available, saving as HTML instead")
                html_path = output_path.replace('.pdf', '.html')
                return self._export_html(results, html_path, metadata)
            
            self.logger.info(f"Exported to PDF: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"PDF export failed: {str(e)}")
            return False
    
    def _export_json(self, results: List[SearchResult], output_path: str,
                    metadata: Optional[dict] = None) -> bool:
        """Export to JSON format"""
        try:
            export_data = {
                "metadata": {
                    "export_date": datetime.now().isoformat(),
                    "total_results": sum(len(r.propositions) for r in results),
                    "version": "4.0.0",
                    **(metadata or {})
                },
                "results": []
            }
            
            for result in results:
                source_data = {
                    "source": result.source.value if result.source else "Unknown",
                    "query": result.query,
                    "search_time": result.search_time,
                    "total_count": result.total_count,
                    "propositions": [prop.to_dict() for prop in result.propositions]
                }
                export_data["results"].append(source_data)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            self.logger.info(f"Exported to JSON: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"JSON export failed: {str(e)}")
            return False
    
    def _export_xlsx(self, results: List[SearchResult], output_path: str,
                    metadata: Optional[dict] = None) -> bool:
        """Export to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Create summary sheet
            summary_sheet = wb.active
            summary_sheet.title = "Resumo"
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            
            # Summary information
            summary_sheet['A1'] = "Monitor de Políticas Públicas - Relatório de Busca"
            summary_sheet['A1'].font = Font(bold=True, size=16)
            summary_sheet.merge_cells('A1:F1')
            
            summary_sheet['A3'] = "Data de Exportação:"
            summary_sheet['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            
            summary_sheet['A4'] = "Total de Resultados:"
            summary_sheet['B4'] = sum(len(r.propositions) for r in results)
            
            if metadata:
                row = 5
                for key, value in metadata.items():
                    summary_sheet[f'A{row}'] = f"{key.title()}:"
                    summary_sheet[f'B{row}'] = str(value)
                    row += 1
            
            # Create sheet for each source
            for result in results:
                if not result.propositions:
                    continue
                
                source_name = result.source.name if result.source else "Unknown"
                sheet_name = source_name[:31]  # Excel sheet name limit
                
                ws = wb.create_sheet(title=sheet_name)
                
                # Headers
                headers = ['Tipo', 'Número', 'Ano', 'Título', 'Resumo', 
                          'Autores', 'Data Publicação', 'Status', 'URL']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Data
                for row, prop in enumerate(result.propositions, 2):
                    ws.cell(row=row, column=1, value=prop.type.value)
                    ws.cell(row=row, column=2, value=prop.number)
                    ws.cell(row=row, column=3, value=prop.year)
                    ws.cell(row=row, column=4, value=prop.title)
                    ws.cell(row=row, column=5, value=prop.summary[:500])
                    ws.cell(row=row, column=6, value=prop.author_names)
                    ws.cell(row=row, column=7, value=prop.publication_date.strftime('%d/%m/%Y'))
                    ws.cell(row=row, column=8, value=prop.status.value)
                    ws.cell(row=row, column=9, value=prop.url)
                
                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            self.logger.info(f"Exported to XLSX: {output_path}")
            return True
            
        except ImportError:
            self.logger.error("openpyxl not installed. Cannot export to XLSX.")
            return False
        except Exception as e:
            self.logger.error(f"XLSX export failed: {str(e)}")
            return False