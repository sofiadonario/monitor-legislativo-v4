"""
API routes for Monitor Legislativo Web
"""

from datetime import datetime
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, Query, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import io
import csv

import sys
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from core.api.api_service import APIService
from core.models.models import SearchResult, APIStatus
from core.utils.export_service import ExportService

router = APIRouter()

# Initialize services
api_service = APIService()
export_service = ExportService()


class SearchRequest(BaseModel):
    """Search request model"""
    query: str
    sources: Optional[List[str]] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    page: Optional[int] = 1
    page_size: Optional[int] = 25


class ExportRequest(BaseModel):
    """Export request model"""
    results: List[Dict[str, Any]]
    format: str = "CSV"
    metadata: Optional[Dict[str, Any]] = None


@router.get("/search")
async def search(
    q: str = Query(..., description="Search query"),
    sources: Optional[str] = Query(None, description="Comma-separated source keys"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(25, ge=1, le=100, description="Results per page")
):
    """
    Search for propositions across multiple sources
    """
    if not q.strip():
        raise HTTPException(status_code=400, detail="Query cannot be empty")
    
    # Parse sources
    source_list = None
    if sources:
        source_list = [s.strip() for s in sources.split(",")]
    
    # Build filters
    filters = {}
    if start_date:
        filters["start_date"] = start_date
    if end_date:
        filters["end_date"] = end_date
    
    try:
        # Perform search
        results = await api_service.search_all(q, filters, source_list)
        
        # Aggregate results
        all_propositions = []
        total_count = 0
        
        for result in results:
            for prop in result.propositions:
                prop_dict = prop.to_dict()
                prop_dict["_source"] = result.source.value if result.source else "Unknown"
                all_propositions.append(prop_dict)
            total_count += result.total_count
        
        # Apply pagination
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_props = all_propositions[start_idx:end_idx]
        
        return {
            "query": q,
            "filters": filters,
            "sources": source_list or list(api_service.get_available_sources().keys()),
            "total_count": total_count,
            "page": page,
            "page_size": page_size,
            "total_pages": (total_count + page_size - 1) // page_size,
            "results": paginated_props
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")


@router.get("/sources")
async def get_sources():
    """
    Get available data sources
    """
    sources = api_service.get_available_sources()
    return {
        "sources": [
            {"key": key, "name": name, "enabled": True}
            for key, name in sources.items()
        ]
    }


@router.get("/status")
async def get_api_status():
    """
    Get current status of all APIs
    """
    try:
        statuses = await api_service.get_api_status()
        
        return {
            "timestamp": datetime.now().isoformat(),
            "services": [
                {
                    "name": status.name,
                    "source": status.source.value,
                    "is_healthy": status.is_healthy,
                    "last_check": status.last_check.isoformat(),
                    "response_time": status.response_time,
                    "error_message": status.error_message
                }
                for status in statuses
            ]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Status check failed: {str(e)}")


@router.post("/export")
async def export_results(request: ExportRequest, background_tasks: BackgroundTasks):
    """
    Export search results to specified format
    """
    # TODO: Implement async export with file upload to storage
    # For now, return a simple response
    
    return {
        "message": "Export functionality is under development",
        "format": request.format,
        "result_count": len(request.results)
    }


@router.post("/export/csv")
async def export_csv(request: ExportRequest):
    """
    Export search results to CSV format
    """
    try:
        from fastapi.responses import StreamingResponse
        import io
        import csv
        
        # Create CSV in memory
        output = io.StringIO()
        fieldnames = [
            'ID', 'Título', 'Resumo', 'Tipo', 'Data', 'Estado', 
            'Município', 'URL', 'Status', 'Autor', 'Câmara', 'Fonte'
        ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for result in request.results:
            writer.writerow({
                'ID': result.get('id', ''),
                'Título': result.get('title', ''),
                'Resumo': result.get('summary', '')[:500],
                'Tipo': result.get('type', ''),
                'Data': result.get('date', ''),
                'Estado': result.get('state', ''),
                'Município': result.get('municipality', ''),
                'URL': result.get('url', ''),
                'Status': result.get('status', ''),
                'Autor': result.get('author', ''),
                'Câmara': result.get('chamber', ''),
                'Fonte': result.get('_source', '')
            })
        
        output.seek(0)
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=monitor_legislativo_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"CSV export failed: {str(e)}")


@router.post("/export/xlsx")
async def export_xlsx(request: ExportRequest):
    """
    Export search results to Excel format
    """
    try:
        from fastapi.responses import StreamingResponse
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados da Busca"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        headers = [
            'ID', 'Título', 'Resumo', 'Tipo', 'Data', 'Estado', 
            'Município', 'URL', 'Status', 'Autor', 'Câmara', 'Fonte'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row, result in enumerate(request.results, 2):
            ws.cell(row=row, column=1, value=result.get('id', ''))
            ws.cell(row=row, column=2, value=result.get('title', ''))
            ws.cell(row=row, column=3, value=result.get('summary', '')[:500])
            ws.cell(row=row, column=4, value=result.get('type', ''))
            ws.cell(row=row, column=5, value=result.get('date', ''))
            ws.cell(row=row, column=6, value=result.get('state', ''))
            ws.cell(row=row, column=7, value=result.get('municipality', ''))
            ws.cell(row=row, column=8, value=result.get('url', ''))
            ws.cell(row=row, column=9, value=result.get('status', ''))
            ws.cell(row=row, column=10, value=result.get('author', ''))
            ws.cell(row=row, column=11, value=result.get('chamber', ''))
            ws.cell(row=row, column=12, value=result.get('_source', ''))
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Add metadata sheet
        ws_meta = wb.create_sheet("Metadados")
        ws_meta.cell(row=1, column=1, value="Data de Exportação")
        ws_meta.cell(row=1, column=2, value=datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        ws_meta.cell(row=2, column=1, value="Total de Resultados")
        ws_meta.cell(row=2, column=2, value=len(request.results))
        
        if request.metadata:
            row = 3
            for key, value in request.metadata.items():
                ws_meta.cell(row=row, column=1, value=key)
                ws_meta.cell(row=row, column=2, value=str(value))
                row += 1
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=monitor_legislativo_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")


@router.delete("/cache")
async def clear_cache(source: Optional[str] = Query(None, description="Specific source to clear")):
    """
    Clear cache for specific source or all sources
    """
    api_service.clear_cache(source)
    
    return {
        "message": f"Cache cleared for {'source: ' + source if source else 'all sources'}",
        "timestamp": datetime.now().isoformat()
    }


@router.get("/proposition/{source}/{id}")
async def get_proposition_details(source: str, id: str):
    """
    Get detailed information about a specific proposition
    """
    # TODO: Implement proposition details endpoint
    raise HTTPException(status_code=501, detail="Not implemented yet")